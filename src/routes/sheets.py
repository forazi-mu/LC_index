from flask import Blueprint, request, jsonify
import openpyxl
import io
import json
from datetime import datetime

sheets_bp = Blueprint('sheets', __name__)

def process_data(data_list: list, header_config: dict = None) -> list:
    """Processes the data list: calculates TOTAL, TOTAL NW, TOTAL GW."""
    for row in data_list:
        # Calculate totals based on existing logic
        if 'QTY/CTN' in row and 'CTNS' in row:
            row['TOTAL'] = row['QTY/CTN'] * row['CTNS']
        if 'CTNS' in row and 'NW' in row:
            row['TOTAL NW'] = row['CTNS'] * row['NW']
        if 'CTNS' in row and 'GW' in row:
            row['TOTAL GW'] = row['CTNS'] * row['GW']
    return data_list

def sort_and_subtotal(data_list: list, header_config: dict = None) -> dict:
    """Sorts by ITEM and calculates subtotals and grand totals."""
    # Sort by ITEM
    data_sorted = sorted(data_list, key=lambda x: x.get('ITEM', ''))
    
    # Prepare result structure
    result = {
        'data': data_sorted,
        'subtotals': {},
        'grand_total': {}
    }
    
    # Group by ITEM and calculate subtotals
    items = {}
    for row in data_sorted:
        item = row.get('ITEM', 'Unknown')
        if item not in items:
            items[item] = []
        items[item].append(row)
    
    # Calculate subtotals for each item
    for item, rows in items.items():
        subtotal = {}
        # Calculate subtotals for numeric fields
        numeric_fields = ['CTNS', 'TOTAL', 'TOTAL NW', 'TOTAL GW', 'QTY/CTN', 'NW', 'GW']
        for field in numeric_fields:
            if any(field in row for row in rows):
                subtotal[field] = sum(row.get(field, 0) for row in rows if isinstance(row.get(field), (int, float)))
        result['subtotals'][item] = subtotal
    
    # Calculate grand total
    grand_total = {}
    numeric_fields = ['CTNS', 'TOTAL', 'TOTAL NW', 'TOTAL GW', 'QTY/CTN', 'NW', 'GW']
    for field in numeric_fields:
        if any(field in row for row in data_sorted):
            grand_total[field] = sum(row.get(field, 0) for row in data_sorted if isinstance(row.get(field), (int, float)))
    result['grand_total'] = grand_total
    
    return result

def parse_excel_file(file_content: bytes, header_config: dict = None) -> list:
    """Parse Excel file content and return list of dictionaries."""
    try:
        # Load workbook from bytes
        workbook = openpyxl.load_workbook(io.BytesIO(file_content))
        worksheet = workbook.active
        
        # Get header row (first row)
        headers = []
        for cell in worksheet[1]:
            if cell.value:
                headers.append(str(cell.value).strip())
        
        # If no headers found, return empty list
        if not headers:
            raise ValueError("No headers found in the Excel file")
        
        # Parse data rows
        data = []
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if not any(row):  # Skip empty rows
                continue
                
            row_data = {}
            for i, header in enumerate(headers):
                if i < len(row) and row[i] is not None:
                    value = row[i]
                    # Convert numeric fields
                    if header in ['QTY/CTN', 'CTNS', 'NW', 'GW'] or 'QTY' in header or 'WEIGHT' in header or 'TOTAL' in header:
                        try:
                            row_data[header] = float(value) if '.' in str(value) else int(value)
                        except (ValueError, TypeError):
                            row_data[header] = 0
                    else:
                        row_data[header] = str(value).strip()
                else:
                    # Set default values for missing data
                    if header in ['QTY/CTN', 'CTNS', 'NW', 'GW'] or 'QTY' in header or 'WEIGHT' in header or 'TOTAL' in header:
                        row_data[header] = 0
                    else:
                        row_data[header] = ''
            
            # Only add rows that have at least some data
            if any(row_data.values()):
                data.append(row_data)
        
        return data
        
    except Exception as e:
        raise ValueError(f"Error parsing Excel file: {str(e)}")

@sheets_bp.route('/headers', methods=['POST'])
def get_file_headers():
    """Extract headers from uploaded Excel file."""
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file uploaded'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        if not file.filename.lower().endswith(('.xlsx', '.xls')):
            return jsonify({'error': 'Please upload an Excel file (.xlsx or .xls)'}), 400
        
        # Read file content
        file_content = file.read()
        
        # Load workbook and get headers
        workbook = openpyxl.load_workbook(io.BytesIO(file_content))
        worksheet = workbook.active
        
        headers = []
        for cell in worksheet[1]:
            if cell.value:
                headers.append(str(cell.value).strip())
        
        return jsonify({'headers': headers})
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@sheets_bp.route('/upload', methods=['POST'])
def upload_excel_file():
    """Handle Excel file upload and process the data."""
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file uploaded'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        if not file.filename.lower().endswith(('.xlsx', '.xls')):
            return jsonify({'error': 'Please upload an Excel file (.xlsx or .xls)'}), 400
        
        # Read file content
        file_content = file.read()
        
        # Get header configuration if provided
        header_config = None
        if 'headers' in request.form:
            try:
                header_config = json.loads(request.form['headers'])
            except:
                pass
        
        # Parse Excel file
        sheets_data = parse_excel_file(file_content, header_config)
        
        if not sheets_data:
            return jsonify({'error': 'No valid data found in Excel file'}), 400
        
        # Process the data
        processed_data = process_data(sheets_data.copy(), header_config)
        result = sort_and_subtotal(processed_data, header_config)
        
        return jsonify(result)
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@sheets_bp.route('/process', methods=['POST'])
def process_sheets_data():
    """Process Google Sheets data with the specified format."""
    try:
        data = request.get_json()
        
        if not data or 'sheets_data' not in data:
            return jsonify({'error': 'No sheets_data provided'}), 400
        
        sheets_data = data['sheets_data']
        
        if not sheets_data:
            return jsonify({'error': 'Empty sheets_data provided'}), 400
        
        # Get header configuration if provided
        header_config = data.get('header_config', None)
        
        # Process the data
        processed_data = process_data(sheets_data.copy(), header_config)
        result = sort_and_subtotal(processed_data, header_config)
        
        return jsonify(result)
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@sheets_bp.route('/sample', methods=['GET'])
def get_sample_data():
    """Get sample data for testing."""
    sample_data = [
        {'ITEM': 'POWER SUPPLY', 'MODEL': 'Power supply', 'QTY/CTN': 10, 'CTNS': 5, 'NW': 18, 'GW': 20},
        {'ITEM': 'POWER SUPPLY', 'MODEL': '12V400W', 'QTY/CTN': 10, 'CTNS': 3, 'NW': 18, 'GW': 20},
        {'ITEM': 'CABLE', 'MODEL': 'HDMI Cable', 'QTY/CTN': 50, 'CTNS': 10, 'NW': 0.5, 'GW': 0.6},
        {'ITEM': 'CABLE', 'MODEL': 'USB Cable', 'QTY/CTN': 100, 'CTNS': 7, 'NW': 0.2, 'GW': 0.25},
        {'ITEM': 'ADAPTER', 'MODEL': 'AC Adapter', 'QTY/CTN': 20, 'CTNS': 4, 'NW': 2.5, 'GW': 3.0},
    ]
    
    processed_data = process_data(sample_data.copy())
    result = sort_and_subtotal(processed_data)
    
    return jsonify(result)

# Document processing routes
@sheets_bp.route('/documents/proforma', methods=['POST'])
def process_proforma_invoice():
    """Process Pro forma Invoice document."""
    try:
        data = request.get_json()
        
        # Validate required fields
        required_fields = ['piNumber', 'piDate']
        for field in required_fields:
            if field not in data or not data[field]:
                return jsonify({'error': f'Missing required field: {field}'}), 400
        
        # Process the document
        result = {
            'document_type': 'Pro forma Invoice',
            'pi_number': data['piNumber'],
            'pi_date': data['piDate'],
            'created_at': datetime.now().isoformat(),
            'status': 'processed'
        }
        
        return jsonify(result)
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@sheets_bp.route('/documents/lc', methods=['POST'])
def process_lc_document():
    """Process LC Document."""
    try:
        data = request.get_json()
        
        # Validate required fields
        required_fields = ['lcNumber', 'lcDate']
        for field in required_fields:
            if field not in data or not data[field]:
                return jsonify({'error': f'Missing required field: {field}'}), 400
        
        # Process the document
        result = {
            'document_type': 'LC Document',
            'lc_number': data['lcNumber'],
            'lc_date': data['lcDate'],
            'created_at': datetime.now().isoformat(),
            'status': 'processed'
        }
        
        return jsonify(result)
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@sheets_bp.route('/documents/packing', methods=['POST'])
def process_packing_list():
    """Process Packing List document."""
    try:
        data = request.get_json()
        
        # Validate required fields
        required_fields = ['ciNumber', 'ciDate', 'lcNumber', 'blNumber']
        for field in required_fields:
            if field not in data or not data[field]:
                return jsonify({'error': f'Missing required field: {field}'}), 400
        
        # Process the document
        result = {
            'document_type': 'Packing List',
            'ci_number': data['ciNumber'],
            'ci_date': data['ciDate'],
            'lc_number': data['lcNumber'],
            'bl_number': data['blNumber'],
            'created_at': datetime.now().isoformat(),
            'status': 'processed'
        }
        
        return jsonify(result)
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@sheets_bp.route('/documents/commercial', methods=['POST'])
def process_commercial_invoice():
    """Process Commercial Invoice document."""
    try:
        data = request.get_json()
        
        # Validate required fields
        required_fields = ['ciNumber', 'ciDate', 'lcNumber', 'blNumber']
        for field in required_fields:
            if field not in data or not data[field]:
                return jsonify({'error': f'Missing required field: {field}'}), 400
        
        # Process the document
        result = {
            'document_type': 'Commercial Invoice',
            'ci_number': data['ciNumber'],
            'ci_date': data['ciDate'],
            'lc_number': data['lcNumber'],
            'bl_number': data['blNumber'],
            'created_at': datetime.now().isoformat(),
            'status': 'processed'
        }
        
        return jsonify(result)
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@sheets_bp.route('/documents/value', methods=['POST'])
def process_value_sheet():
    """Process Value Sheet document."""
    try:
        data = request.get_json()
        
        # Validate required fields
        required_fields = ['lcNumber', 'beNumber', 'beDate']
        for field in required_fields:
            if field not in data or not data[field]:
                return jsonify({'error': f'Missing required field: {field}'}), 400
        
        # Process the document
        result = {
            'document_type': 'Value Sheet',
            'lc_number': data['lcNumber'],
            'be_number': data['beNumber'],
            'be_date': data['beDate'],
            'created_at': datetime.now().isoformat(),
            'status': 'processed'
        }
        
        return jsonify(result)
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@sheets_bp.route('/documents/loading', methods=['POST'])
def process_loading_list():
    """Process Loading List document."""
    try:
        data = request.get_json()
        
        # Validate required fields
        required_fields = ['lcNumber', 'loadingDate', 'loadingPort', 'destinationPort', 'containerNumber', 'totalCarton', 'totalWeight']
        for field in required_fields:
            if field not in data or not data[field]:
                return jsonify({'error': f'Missing required field: {field}'}), 400
        
        # Process the document
        result = {
            'document_type': 'Loading List',
            'lc_number': data['lcNumber'],
            'loading_date': data['loadingDate'],
            'loading_port': data['loadingPort'],
            'destination_port': data['destinationPort'],
            'container_number': data['containerNumber'],
            'total_carton': data['totalCarton'],
            'total_weight': data['totalWeight'],
            'created_at': datetime.now().isoformat(),
            'status': 'processed'
        }
        
        return jsonify(result)
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# Document management routes
@sheets_bp.route('/document-types', methods=['GET'])
def get_document_types():
    """Get list of available document types."""
    document_types = [
        {
            'id': 'proforma',
            'name': 'Pro forma Invoice',
            'description': 'Generate proforma invoices',
            'fields': ['PI Number', 'PI Date'],
            'active': True
        },
        {
            'id': 'lc',
            'name': 'LC Document',
            'description': 'Letter of Credit processing',
            'fields': ['LC Number', 'LC Date'],
            'active': True
        },
        {
            'id': 'packing',
            'name': 'Packing List',
            'description': 'Generate packing lists',
            'fields': ['Commercial Invoice Number', 'Commercial Invoice Date', 'LC Number', 'B/L Number'],
            'active': True
        },
        {
            'id': 'commercial',
            'name': 'Commercial Invoice',
            'description': 'Create commercial invoices',
            'fields': ['Commercial Invoice Number', 'Commercial Invoice Date', 'LC Number', 'B/L Number'],
            'active': True
        },
        {
            'id': 'value',
            'name': 'Value Sheet',
            'description': 'Generate value sheets',
            'fields': ['LC Number', 'Bill of Entry Number', 'B/E Date'],
            'active': True
        },
        {
            'id': 'loading',
            'name': 'Loading List',
            'description': 'Create loading lists',
            'fields': ['LC Number', 'Loading Date', 'Loading Port', 'Destination Port', 'Container Number', 'Total Carton', 'Total Weight'],
            'active': True
        }
    ]
    
    return jsonify(document_types)

@sheets_bp.route('/document-types', methods=['POST'])
def add_document_type():
    """Add new document type."""
    try:
        data = request.get_json()
        
        # Validate required fields
        required_fields = ['name', 'description', 'fields']
        for field in required_fields:
            if field not in data or not data[field]:
                return jsonify({'error': f'Missing required field: {field}'}), 400
        
        # Create new document type
        new_doc_type = {
            'id': data['name'].lower().replace(' ', '_'),
            'name': data['name'],
            'description': data['description'],
            'fields': data['fields'].split(',') if isinstance(data['fields'], str) else data['fields'],
            'active': True,
            'created_at': datetime.now().isoformat()
        }
        
        return jsonify(new_doc_type)
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

